from __future__ import annotations

import hashlib
import importlib.util
from pathlib import Path
from typing import Any

from enterprise_energy_research.adapters.base import AdapterHealth, ArtifactResult
from enterprise_energy_research.domain.enums import ArtifactType
from enterprise_energy_research.domain.models import ArtifactBinding, FrozenResearchBundle
from enterprise_energy_research.vendor import embedded_skill_root


class ExcelMasterFrozenPublisher:
    name = "excel_master"
    artifact_type = ArtifactType.EXCEL

    def __init__(self, skill_root: Path | None = None, *, theme: str = "deep-navy", freeze_rows: int = 1) -> None:
        self.skill_root = skill_root or embedded_skill_root("excel-master")
        self.theme = theme
        self.freeze_rows = freeze_rows

    def health(self) -> AdapterHealth:
        available = (self.skill_root / "scripts" / "make_excel.py").is_file() and importlib.util.find_spec("pandas") is not None and importlib.util.find_spec("openpyxl") is not None
        return AdapterHealth(name=self.name, available=available, version="excel-master", diagnostics=[] if available else ["Excel Master、pandas 或 openpyxl 不可用"])

    def publish(self, bundle: FrozenResearchBundle, binding: ArtifactBinding, output_path: Path) -> ArtifactResult:
        health = self.health()
        if not health.available:
            return ArtifactResult(adapter=self.name, artifact_id=binding.artifact_id, artifact_type=binding.type, status="failed", diagnostics=health.diagnostics)
        if binding.type != self.artifact_type:
            return ArtifactResult(adapter=self.name, artifact_id=binding.artifact_id, artifact_type=binding.type, status="failed", diagnostics=["Excel 发布器收到非 Excel 绑定"])
        import pandas as pd
        module_path = self.skill_root / "scripts" / "make_excel.py"
        spec = importlib.util.spec_from_file_location("eer_excel_master", module_path)
        if spec is None or spec.loader is None:
            raise RuntimeError("无法加载 Excel Master")
        module = importlib.util.module_from_spec(spec); spec.loader.exec_module(module)
        canonical_id = bundle.run_manifest.canonical_entity_id
        entities = [self._row(item) for item in bundle.entities]
        canonical_rows = [row for row in entities if row.get("entity_id") == canonical_id]
        for row in canonical_rows:
            # Keep provenance inside the fixed 17-sheet workbook contract.
            # Adding an 18th legacy metadata sheet would violate the Skill's
            # required worksheet set.
            row["run_id"] = bundle.run_manifest.run_id
            row["freeze_id"] = bundle.freeze.freeze_id
            row["root_hash"] = bundle.freeze.root_hash
        products = [self._row(item) for item in bundle.products]
        product_parameters = [
            {
                "product_id": product.product_id,
                "entity_id": product.entity_id,
                "product_name": product.name,
                "parameter_name": parameter.name,
                "value": parameter.value,
                "unit": parameter.unit,
                "claim_ids": str(parameter.claim_ids),
            }
            for product in bundle.products for parameter in product.parameters
        ]
        numeric_claims = [
            self._row(claim) for claim in bundle.claims
            if isinstance(claim.value, (int, float)) and not isinstance(claim.value, bool)
        ]
        solution_rows = [self._row(item) for item in bundle.solutions]

        def solutions(*engines: str):
            return pd.DataFrame([row for row in solution_rows if row.get("engine") in engines])

        # The workbook contract is fixed even when a section has no rows: an
        # empty, named sheet is an honest gap; silently omitting the sheet is
        # not a valid Excel Master delivery.
        sheets = [
            ("01_企业基本信息", pd.DataFrame(canonical_rows)),
            ("02_集团及子公司", pd.DataFrame([row for row in entities if row.get("entity_id") != canonical_id])),
            ("03_生产基地", pd.DataFrame([self._row(item) for item in bundle.factories])),
            ("04_产品矩阵", pd.DataFrame(products)),
            ("05_产品参数", pd.DataFrame(product_parameters)),
            ("06_经营数据", pd.DataFrame([self._row(item) for item in bundle.claims if item.field_name in {
                "revenue", "profit", "gross_margin", "rnd_expense", "rnd_expense_ratio",
                "operating_cash_flow", "employee_count", "investment", "business_segment",
            }])),
            ("07_工艺与用能", pd.DataFrame([
                *[self._row(item) for item in bundle.energy_profiles],
                *[self._row(item) for item in bundle.claims if item.field_name in {
                    "process", "processes", "production_lines", "energy_consumption",
                    "electricity_consumption", "energy_equipment", "electricity_load",
                    "roof_area", "transformer_capacity",
                }],
            ])),
            ("08_EPC机会", solutions("EPC", "PV_EPC")),
            ("09_零碳节能", solutions("ZERO_CARBON", "ENERGY_EFFICIENCY", "GREEN_POWER", "ZERO_CARBON_FACTORY")),
            ("10_储能ODM", solutions("STORAGE_ODM", "STORAGE", "ODM", "V2G")),
            ("11_出海合作", solutions("OVERSEAS", "CHANNEL")),
            ("12_原始事实", pd.DataFrame([self._row(item) for item in bundle.claims])),
            ("13_来源URL", pd.DataFrame([self._row(item) for item in bundle.sources])),
            ("14_图片来源", pd.DataFrame([self._row(item) for item in bundle.images])),
            ("15_冲突数据", pd.DataFrame([self._row(item) for item in bundle.conflicts])),
            ("16_数据缺口", pd.DataFrame([self._row(item) for item in bundle.gaps])),
            ("17_图表数据", pd.DataFrame(numeric_claims)),
        ]
        sheets = [
            (name, frame if not frame.empty else pd.DataFrame([{"数据状态": "暂无已核验数据"}]))
            for name, frame in sheets
        ]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        module.make_excel(sheets, output_path, theme=self.theme, freeze_rows=self.freeze_rows)
        self._apply_layout(output_path)
        digest = hashlib.sha256(output_path.read_bytes()).hexdigest()
        return ArtifactResult(adapter=self.name, artifact_id=binding.artifact_id, artifact_type=binding.type, path=output_path, content_sha256=digest, used_claim_ids=list(binding.claim_ids), used_image_ids=list(binding.image_ids), status="published")

    @staticmethod
    def _row(model: Any) -> dict[str, Any]:
        raw = model.model_dump(mode="json")
        return {key: value if isinstance(value, (str, int, float, bool)) or value is None else str(value) for key, value in raw.items()}

    @staticmethod
    def _apply_layout(output_path: Path) -> None:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        widths = {
            "claim_id": 20, "entity_id": 20, "factory_id": 20, "product_id": 20,
            "source_id": 18, "solution_id": 20, "energy_profile_id": 20, "gap_id": 20,
            "canonical_name": 28, "canonical_company_name": 28, "name": 26, "address": 24,
            "field_name": 30, "value": 34, "raw_text": 42, "context_text": 34,
            "opportunity": 44, "proposed_solution": 44, "benefit_logic": 38,
            "data_requirements": 32, "risks": 30, "next_step": 36, "assumptions": 38,
            "processes": 34, "parameters": 34, "description": 34, "canonical_url": 44,
            "source_title": 34, "scope": 30, "qualifier": 16, "notes": 28,
        }
        workbook = load_workbook(output_path)
        try:
            for sheet in workbook.worksheets:
                sheet.sheet_view.zoomScale = 85
                sheet.freeze_panes = "B2"
                sheet.auto_filter.ref = sheet.dimensions if sheet.max_row > 1 else None
                for column in range(2, sheet.max_column + 1):
                    header = str(sheet.cell(1, column).value or "")
                    if header in widths:
                        sheet.column_dimensions[get_column_letter(column)].width = widths[header]
                for row in sheet.iter_rows(min_row=2):
                    long_row = False
                    for cell in row[1:]:
                        text = "" if cell.value is None else str(cell.value)
                        if len(text) > 24:
                            long_row = True
                        cell.alignment = Alignment(
                            horizontal="right" if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool) else "left",
                            vertical="top", wrap_text=True,
                        )
                    if long_row:
                        sheet.row_dimensions[row[0].row].height = 36
                sheet.row_dimensions[1].height = 28
            workbook.save(output_path)
        finally:
            workbook.close()
