from __future__ import annotations

import hashlib
import json
import zipfile
from pathlib import Path

from pydantic import BaseModel, Field

from enterprise_energy_research.adapters.base import ArtifactResult
from enterprise_energy_research.domain.enums import ArtifactStatus, ArtifactType, ValidationStatus
from enterprise_energy_research.domain.models import ArtifactManifest, FrozenResearchBundle
from enterprise_energy_research.validation.delivery_quality import (
    PptVisualDeliveryRecord,
    inspect_ppt_visual_delivery,
    inspect_word_depth,
)


class ConsistencyFinding(BaseModel):
    code: str
    artifact_id: str | None = None
    message: str


class ConsistencyReport(BaseModel):
    freeze_id: str
    status: ValidationStatus
    findings: list[ConsistencyFinding] = Field(default_factory=list)


class ArtifactConsistencyAuditor:
    """Parse generated artifacts and verify freeze, checksum and binding consistency."""

    def audit(
        self,
        bundle: FrozenResearchBundle,
        manifest: ArtifactManifest,
        results: list[ArtifactResult],
    ) -> ConsistencyReport:
        findings: list[ConsistencyFinding] = []
        if manifest.freeze_id != bundle.freeze.freeze_id:
            findings.append(ConsistencyFinding(code="MANIFEST_FREEZE_MISMATCH", message="Artifact manifest does not belong to the frozen bundle"))
        results_by_id = {item.artifact_id: item for item in results}
        for binding in manifest.artifacts:
            result = results_by_id.get(binding.artifact_id)
            if not result:
                findings.append(ConsistencyFinding(code="MISSING_ARTIFACT_RESULT", artifact_id=binding.artifact_id, message="No publisher result exists"))
                continue
            if binding.status == ArtifactStatus.SKIPPED:
                if result.status != "skipped":
                    findings.append(ConsistencyFinding(code="SKIP_STATUS_MISMATCH", artifact_id=binding.artifact_id, message="Manifest skip decision was not preserved"))
                continue
            if result.status != "published" or not result.path:
                findings.append(ConsistencyFinding(code="ARTIFACT_NOT_PUBLISHED", artifact_id=binding.artifact_id, message="Planned artifact was not published"))
                continue
            path = Path(result.path)
            if not path.is_file():
                findings.append(ConsistencyFinding(code="ARTIFACT_FILE_MISSING", artifact_id=binding.artifact_id, message=str(path)))
                continue
            digest = hashlib.sha256(path.read_bytes()).hexdigest()
            if digest != result.content_sha256:
                findings.append(ConsistencyFinding(code="ARTIFACT_HASH_MISMATCH", artifact_id=binding.artifact_id, message="Artifact changed after publisher checksum"))
            if not set(result.used_claim_ids).issubset(binding.claim_ids):
                findings.append(ConsistencyFinding(code="UNBOUND_CLAIM_USED", artifact_id=binding.artifact_id, message="Publisher used a claim outside the artifact binding"))
            if not set(result.used_image_ids).issubset(binding.image_ids):
                findings.append(ConsistencyFinding(code="UNBOUND_IMAGE_USED", artifact_id=binding.artifact_id, message="Publisher used an image outside the artifact binding"))
            self._parse_artifact(path, binding.type, bundle, binding.artifact_id, findings)
        status = ValidationStatus.BLOCKED if findings else ValidationStatus.PASS
        return ConsistencyReport(freeze_id=bundle.freeze.freeze_id, status=status, findings=findings)

    @staticmethod
    def _parse_artifact(
        path: Path,
        kind: ArtifactType,
        bundle: FrozenResearchBundle,
        artifact_id: str,
        findings: list[ConsistencyFinding],
    ) -> None:
        try:
            fixture_mode = bundle.run_manifest.model_gateway.get("mode") in {
                "fixture", "recorded-fixture", "recorded-fixture-only",
            }
            if kind in {ArtifactType.ENTERPRISE_HTML, ArtifactType.PRODUCT_HTML}:
                text = path.read_text(encoding="utf-8")
                if bundle.freeze.root_hash not in text or bundle.freeze.freeze_id not in text:
                    raise ValueError("HTML does not embed freeze provenance")
            elif kind == ArtifactType.WORD:
                with zipfile.ZipFile(path) as archive:
                    text = "".join(
                        archive.read(name).decode("utf-8", errors="ignore")
                        for name in archive.namelist() if name.startswith("word/") and name.endswith(".xml")
                    )
                if bundle.freeze.freeze_id not in text:
                    raise ValueError("Word report does not embed freeze ID")
                if not fixture_mode:
                    rendered_pdf = path.with_suffix(".pdf")
                    depth = inspect_word_depth(path, rendered_pdf=rendered_pdf)
                    if depth.status != "PASS":
                        raise ValueError("Word formal-depth gate failed: " + "; ".join(depth.findings))
            elif kind == ArtifactType.EXCEL:
                from openpyxl import load_workbook
                workbook = load_workbook(path, read_only=True, data_only=False)
                try:
                    values: dict[str, object] = {}
                    if "运行清单" in workbook.sheetnames:
                        # Backward-compatible reader for historical releases.
                        for row in workbook["运行清单"].iter_rows(values_only=True):
                            compact = [value for value in row if value is not None]
                            if len(compact) >= 2 and compact[0] not in {"field", "运行清单"}:
                                values[str(compact[0])] = compact[1]
                    elif "01_企业基本信息" in workbook.sheetnames:
                        # Current Skill contract has exactly 17 business
                        # sheets.  Provenance is stored as columns on the
                        # canonical enterprise row instead of a hidden or
                        # ad-hoc eighteenth sheet.
                        rows = workbook["01_企业基本信息"].iter_rows(values_only=True)
                        headers = [str(value or "") for value in next(rows, ())]
                        first = next(rows, ())
                        values = {header: value for header, value in zip(headers, first) if header}
                    else:
                        raise ValueError("Excel workbook has no canonical enterprise provenance sheet")
                    if values.get("freeze_id") != bundle.freeze.freeze_id or values.get("root_hash") != bundle.freeze.root_hash:
                        raise ValueError("Excel run manifest provenance mismatch")
                finally:
                    workbook.close()
            elif kind == ArtifactType.PPT:
                with zipfile.ZipFile(path) as archive:
                    slide_count = len([name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")])
                    embedded_media_count = len([
                        name for name in archive.namelist()
                        if name.startswith("ppt/media/") and not name.endswith("/")
                    ])
                    all_xml = "".join(
                        archive.read(name).decode("utf-8", errors="ignore")
                        for name in archive.namelist() if name.endswith(".xml")
                    )
                if not 15 <= slide_count <= 20:
                    raise ValueError(f"PPT slide count {slide_count} is outside 15-20")
                if bundle.freeze.freeze_id not in all_xml:
                    raise ValueError("PPT does not embed freeze ID")
                if not fixture_mode:
                    quality_path = Path(str(path) + ".quality.json")
                    if not quality_path.is_file():
                        raise ValueError(f"PPT visual quality record is missing: {quality_path.name}")
                    record = PptVisualDeliveryRecord.model_validate(json.loads(quality_path.read_text(encoding="utf-8")))
                    evidence_map_path = path.parent / f"{path.stem}_ppt_master_project" / "presentation_evidence_map.json"
                    if not evidence_map_path.is_file():
                        raise ValueError("PPT presentation evidence map is missing beside the final deck")
                    evidence_map = json.loads(evidence_map_path.read_text(encoding="utf-8"))
                    expected_image_count = len(evidence_map.get("required_verified_image_ids", []))
                    if record.required_verified_image_count != expected_image_count:
                        raise ValueError(
                            f"PPT quality record declares {record.required_verified_image_count} required images; evidence map requires {expected_image_count}"
                        )
                    if record.embedded_verified_image_count != expected_image_count or embedded_media_count < expected_image_count:
                        raise ValueError("PPTX package does not contain every contracted verified evidence image")
                    visual_findings = inspect_ppt_visual_delivery(record)
                    if visual_findings:
                        raise ValueError("PPT visual-delivery gate failed: " + "; ".join(visual_findings))
        except Exception as exc:
            findings.append(ConsistencyFinding(code="ARTIFACT_PARSE_FAILED", artifact_id=artifact_id, message=f"{kind.value}: {exc}"))
